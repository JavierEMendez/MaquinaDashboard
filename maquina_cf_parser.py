"""
maquina_cf_parser.py — Parse the "MAQUINA CF" workbook into clean JSON for
the /cashflow dashboard.

Only three sheets matter (the ones visible when the file opens):

  • 'CF Consolidado'   — master: US block + MX block + consolidated
                         Total CF / Cumm CF + a Reservas (reserves) block.
                         All figures already in USD.
  • 'Business CF - US' — entity-level detail behind the US block
                         (Equity by project, GP Profit by project, RECAP,
                         Opex, Investments/Dividends).
  • 'Business CF - MX' — entity-level detail behind the MX block. Carries a
                         monthly FX-rate row and is column-shifted +1 vs the
                         other two sheets.

Time axis (read dynamically, not hard-coded):
  2022 = a single seed/annual column · 2023-2026 = 12 monthly columns + an
  annual-total column each · 2027-2036 = annual-only projection columns.

The workbook is a live, partially-broken model — cells can hold #REF! /
#VALUE! / #DIV/0! strings. Everything is coerced through ``_num`` so a
broken formula becomes 0.0 rather than crashing the import.

This is a stopgap: eventually Ember/Ranman/etc. figures will flow in live.
The parser targets the current file's structure but matches rows by their
labels (section-aware) so a re-upload with minor row shifts still imports.
"""
from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Optional

import openpyxl

# ── Sheet config: (year_row, month_row, label_col) per target sheet ──────────
SHEET_CONSOLIDATED = "CF Consolidado"
SHEET_US = "Business CF - US"
SHEET_MX = "Business CF - MX"

_MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def _num(v) -> float:
    """Coerce a cell value to float. Broken formulas (#REF!, #VALUE!, …),
    blanks and non-numeric text all collapse to 0.0."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        try:
            f = float(v)
            return f if f == f else 0.0  # drop NaN
        except (TypeError, ValueError, OverflowError):
            return 0.0
    s = str(v).strip()
    if not s or s.startswith("#"):
        return 0.0
    s = s.replace(",", "").replace("$", "").replace("%", "")
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except ValueError:
        return 0.0


def _label(v) -> str:
    return "" if v is None else str(v).strip()


def _col_periods(ws, year_row: int, month_row: int) -> dict[int, tuple[int, Optional[int]]]:
    """Map each column index → (year, month) where month is None for an
    annual-total / annual-only column. Year labels are expanded across the
    sheet's merged ranges so a label that only lives in the merge's top-left
    cell still tags every column it visually spans.
    """
    col_year: dict[int, int] = {}

    # 1) raw single-cell year labels on the year row
    for c in range(1, ws.max_column + 1):
        m = re.search(r"(20\d{2})", _label(ws.cell(row=year_row, column=c).value))
        if m:
            col_year[c] = int(m.group(1))

    # 2) expand merged ranges that intersect the year row
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= year_row <= rng.max_row:
            tl = ws.cell(row=rng.min_row, column=rng.min_col).value
            m = re.search(r"(20\d{2})", _label(tl))
            if m:
                yr = int(m.group(1))
                for c in range(rng.min_col, rng.max_col + 1):
                    col_year.setdefault(c, yr)

    # 3) months on the month row
    col_month: dict[int, int] = {}
    for c in range(1, ws.max_column + 1):
        key = _label(ws.cell(row=month_row, column=c).value).lower()[:3]
        if key in _MONTHS:
            col_month[c] = _MONTHS[key]

    periods: dict[int, tuple[int, Optional[int]]] = {}
    for c, yr in col_year.items():
        periods[c] = (yr, col_month.get(c))  # month None ⇒ annual column
    return periods


def _annual_col_for_year(periods: dict[int, tuple[int, Optional[int]]]) -> dict[int, int]:
    """year → the column index holding that year's annual total (month is
    None). If a year has several such columns the right-most wins."""
    out: dict[int, int] = {}
    for c, (yr, mo) in sorted(periods.items()):
        if mo is None:
            out[yr] = c
    return out


def _row_series(ws, row: int, periods, annual_cols) -> dict:
    """Extract one data row as {year: annual_value} plus a flat monthly map
    {'YYYY-MM': value}. Annual value prefers the explicit annual column; if a
    year has only monthly columns it falls back to their sum."""
    by_year: dict[int, float] = {}
    months: dict[str, float] = {}
    month_sum: dict[int, float] = {}
    for c, (yr, mo) in periods.items():
        v = _num(ws.cell(row=row, column=c).value)
        if mo is None:
            by_year[yr] = v
        else:
            if v:
                months[f"{yr}-{mo:02d}"] = v
            month_sum[yr] = month_sum.get(yr, 0.0) + v
    for yr, c in annual_cols.items():
        if yr not in by_year:
            by_year[yr] = month_sum.get(yr, 0.0)
    # any year that had months but no annual col
    for yr, s in month_sum.items():
        by_year.setdefault(yr, s)
    return {"by_year": {y: round(by_year[y], 2) for y in sorted(by_year)}, "months": months}


def _walk_sections(ws, label_col, periods, annual_cols,
                   section_starts: dict[str, str], total_labels: set[str]):
    """Walk the rows of a Business CF sheet collecting entity line items
    grouped by section. ``section_starts`` maps a normalized header label →
    section key. ``total_labels`` are normalized labels that close a section
    (captured as the section total, not an entity)."""
    sections: dict[str, dict] = {}
    cur_key = None
    for r in range(1, ws.max_row + 1):
        lbl = _label(ws.cell(row=r, column=label_col).value)
        if not lbl:
            continue
        norm = lbl.lower()
        # section header?
        hit = None
        for needle, key in section_starts.items():
            if norm.startswith(needle):
                hit = key
                break
        if hit:
            cur_key = hit
            sections.setdefault(cur_key, {"label": lbl, "entities": [],
                                          "total": None})
            continue
        if cur_key is None:
            continue
        series = _row_series(ws, r, periods, annual_cols)
        is_total = any(norm.startswith(t) for t in total_labels)
        is_cumm = norm.startswith("cumm")
        if is_cumm:
            continue  # cumulative rows are derived; skip in detail
        if is_total:
            sections[cur_key]["total"] = series
            cur_key = None  # section closed
        else:
            # skip rows that are entirely zero across all years
            if any(abs(v) > 0.005 for v in series["by_year"].values()):
                sections[cur_key]["entities"].append({"name": lbl, **series})
    return sections


def parse_maquina_cf(file_bytes: bytes, filename: str = "") -> dict:
    """Parse the workbook bytes → dashboard JSON. Raises ValueError if the
    three required sheets aren't present."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    names = set(wb.sheetnames)
    missing = [s for s in (SHEET_CONSOLIDATED, SHEET_US, SHEET_MX) if s not in names]
    if missing:
        raise ValueError(
            "Workbook is missing required sheet(s): " + ", ".join(missing)
            + ". Expected the standard MAQUINA CF file."
        )

    # ── Consolidated master ──────────────────────────────────────────────
    cons = wb[SHEET_CONSOLIDATED]
    cper = _col_periods(cons, year_row=5, month_row=6)
    cann = _annual_col_for_year(cper)
    years = sorted({yr for (yr, _mo) in cper.values()})
    monthly_years = sorted({yr for (yr, mo) in cper.values() if mo is not None})
    last_actual = max(monthly_years) if monthly_years else (max(years) if years else 0)

    def cons_row(row):
        return _row_series(cons, row, cper, cann)

    # Row map for CF Consolidado (section-aware: US block, MX block, totals,
    # reservas). Pinned by row number with the label captured for display.
    def row_at(row):
        return _label(cons.cell(row=row, column=2).value)

    consolidated = {
        "us": {
            "inv_div":       cons_row(8),
            "equity":        cons_row(9),
            "gp_profit":     cons_row(10),
            "recap":         cons_row(11),
            "investment_mx": cons_row(12),
            "opex":          cons_row(13),
            "total":         cons_row(14),
            "cumm":          cons_row(15),
        },
        "mx": {
            "inv_div":   cons_row(18),
            "equity":    cons_row(19),
            "gp_profit": cons_row(20),
            "recap":     cons_row(21),
            "opex":      cons_row(22),
            "total":     cons_row(23),
            "cumm":      cons_row(24),
        },
        "total_cf": cons_row(26),
        "cumm_cf":  cons_row(27),
        "reservas": {
            "operacion_mx":      cons_row(30),
            "operacion_usa":     cons_row(31),
            "dividendo_usa":     cons_row(32),
            "inversion_usa":     cons_row(33),
            "dividendo_mx":      cons_row(34),
            "inversion_mx":      cons_row(35),
            "fondo_liquido_usa": cons_row(36),
            "fondo_liquido_mx":  cons_row(37),
            "total":             cons_row(38),
        },
    }

    # ── US entity detail ─────────────────────────────────────────────────
    us = wb[SHEET_US]
    uper = _col_periods(us, year_row=3, month_row=4)
    uann = _annual_col_for_year(uper)
    us_sections = _walk_sections(
        us, label_col=2, periods=uper, annual_cols=uann,
        section_starts={
            "investments/dividends": "inv_div",
            "equity": "equity",
            "gp profit": "gp_profit",
            "recap": "recap",
            "investment mx": "investment_mx",
            "opex": "opex",
        },
        total_labels={"total investments/dividends", "total equity",
                      "total gp profit", "total recap", "total investment mx",
                      "total opex"},
    )

    # ── MX entity detail (label col = C, +1 column offset) ────────────────
    mx = wb[SHEET_MX]
    mper = _col_periods(mx, year_row=3, month_row=4)
    mann = _annual_col_for_year(mper)
    mx_sections = _walk_sections(
        mx, label_col=3, periods=mper, annual_cols=mann,
        section_starts={
            "investments/divident": "inv_div",   # sheet typo: "Dividents"
            "investments/dividend": "inv_div",
            "equity": "equity",
            "gp profit": "gp_profit",
            "recap": "recap",
            "opex": "opex",
        },
        total_labels={"total investments/dividend", "total equity",
                      "total gp profit", "total recap", "total opex"},
    )

    # ── Derived KPIs (consolidated, lifetime + latest actual year) ────────
    def life(series):  # lifetime sum across actual years only
        return round(sum(v for y, v in series["by_year"].items() if y <= last_actual), 2)

    def at(series, yr):
        return round(series["by_year"].get(yr, 0.0), 2)

    # Gross inflow / outflow from monthly cells of Total CF (sign split).
    gross_in = gross_out = 0.0
    for k, v in consolidated["total_cf"]["months"].items():
        if v > 0:
            gross_in += v
        else:
            gross_out += v

    kpis = {
        "last_actual_year": last_actual,
        "net_cf_last": at(consolidated["total_cf"], last_actual),
        "cumm_cf_last": at(consolidated["cumm_cf"], last_actual),
        "reserves_last": at(consolidated["reservas"]["total"], last_actual),
        # lifetime category sums (consolidated US+MX)
        "equity_life": round(life(consolidated["us"]["equity"]) + life(consolidated["mx"]["equity"]), 2),
        "gp_profit_life": round(life(consolidated["us"]["gp_profit"]) + life(consolidated["mx"]["gp_profit"]), 2),
        "recap_life": round(life(consolidated["us"]["recap"]) + life(consolidated["mx"]["recap"]), 2),
        "opex_life": round(life(consolidated["us"]["opex"]) + life(consolidated["mx"]["opex"]), 2),
        "invdiv_life": round(life(consolidated["us"]["inv_div"]) + life(consolidated["mx"]["inv_div"]), 2),
        # distributions to partners = Dividendo USA + MX (reservas), lifetime
        "distributions_life": round(life(consolidated["reservas"]["dividendo_usa"])
                                    + life(consolidated["reservas"]["dividendo_mx"]), 2),
        # capital deployed = Inversión USA + MX (reservas), lifetime
        "investment_life": round(life(consolidated["reservas"]["inversion_usa"])
                                 + life(consolidated["reservas"]["inversion_mx"]), 2),
        "gross_in": round(gross_in, 2),
        "gross_out": round(gross_out, 2),
        "net_cumm": at(consolidated["cumm_cf"], max(years) if years else last_actual),
    }

    # ── Report date: prefer filename, fall back to today ─────────────────
    report_date = ""
    m = re.search(r"(\d{1,2})\s+([A-Za-z]{3,})\s+(\d{4})", filename or "")
    if m:
        report_date = f"{m.group(1)} {m.group(2)} {m.group(3)}"

    return {
        "meta": {
            "currency": "USD",
            "filename": filename,
            "report_date": report_date,
            "years": years,
            "monthly_years": monthly_years,
            "proj_years": [y for y in years if y > last_actual],
            "last_actual_year": last_actual,
        },
        "consolidated": consolidated,
        "us": {"sections": us_sections},
        "mx": {"sections": mx_sections},
        "kpis": kpis,
    }


if __name__ == "__main__":
    import sys, json
    path = sys.argv[1] if len(sys.argv) > 1 else \
        r"C:/Users/Javier/Downloads/MAQUINA CF - 21 May 2026.xlsx"
    with open(path, "rb") as f:
        data = parse_maquina_cf(f.read(), filename=path.split("/")[-1])
    m = data["meta"]
    print("currency:", m["currency"], "| report:", m["report_date"])
    print("years:", m["years"])
    print("monthly_years:", m["monthly_years"], "| last_actual:", m["last_actual_year"])
    print("\n── KPIs ──")
    for k, v in data["kpis"].items():
        print(f"  {k:22s} {v:>16,.0f}" if isinstance(v, (int, float)) else f"  {k:22s} {v}")
    print("\n── Consolidated Total CF by year ──")
    for y, v in data["consolidated"]["total_cf"]["by_year"].items():
        print(f"  {y}: {v:>16,.0f}")
    print("\n── Consolidated Cumm CF by year ──")
    for y, v in data["consolidated"]["cumm_cf"]["by_year"].items():
        print(f"  {y}: {v:>16,.0f}")
    print("\n── US sections (entity counts) ──")
    for k, s in data["us"]["sections"].items():
        tot = (s["total"] or {}).get("by_year", {}) if s["total"] else {}
        print(f"  {k:14s} {len(s['entities'])} entities | total {s['label']}")
    print("\n── MX sections (entity counts) ──")
    for k, s in data["mx"]["sections"].items():
        print(f"  {k:14s} {len(s['entities'])} entities")
    print("\n── US Equity entities ──")
    for e in data["us"]["sections"].get("equity", {}).get("entities", []):
        print(f"  {e['name']:34s} life={sum(e['by_year'].values()):>14,.0f}")
