"""
polaris_parser.py — Parse Valoran's "Modelo Polaris" workbook (META / IPC toll
roads) into the JSON the Meta company page renders.

What we take from the model, and from where:

  • 'Ingresos META' / 'Ingresos IPC' — monthly toll revenue per highway
    (nominal MXN, sin IVA). Row 1 carries the monthly date axis; the third
    label block ("Ingresos") lists each highway followed by Auto/Bus/CU/CA1/
    CA2 rows and a "Total" row — we take the Total. Sheet-level lifetime
    totals (col C) are used to cross-check the parse.
  • 'Saldos Iniciales' — the "Datos de Ingresos" inputs block: opening TPDA
    per vehicle class and tarifas con / sin IVA per highway.
  • 'Premisas' — model start, concession end dates.
  • 'Portada' — the model's own real discount rate ("T.D Real").
  • 'Flujo META Anual' / 'Flujo IPC Anual' — lifetime EBITDA margin per
    company (Σ EBITDA / Σ Ingresos), used to turn revenue into an
    enterprise-value proxy on the EV Analysis tab.

Rows are matched by label (accent- and case-insensitive), not by number, so a
re-upload with modest row shifts still parses. Broken formula cells (#REF!,
#DIV/0!) coerce to 0.0 rather than crashing the import.

The canonical monthly axis starts at the model start and ends at the last
month in which ANY highway earns revenue — Dec-2050 in the Aug-2026 file,
even though the IPC concession runs to Oct-2064. We deliberately do not
extrapolate past the model; the UI flags the horizon instead.
"""
from __future__ import annotations

import datetime as dt
import io
import re
import unicodedata

import openpyxl

CLASSES = ["Auto", "Bus", "CU", "CA1", "CA2"]
_CLASS_BY_NORM = {c.lower(): c for c in CLASSES}


# ── helpers ──────────────────────────────────────────────────────────────────
def _norm(s) -> str:
    s = "" if s is None else str(s)
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", s).strip().lower()


def slug(s) -> str:
    return re.sub(r"[^a-z0-9]+", "_", _norm(s)).strip("_")


def _num(v) -> float:
    if isinstance(v, bool):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    return 0.0


def _ym(v):
    if isinstance(v, (dt.datetime, dt.date)):
        return "%04d-%02d" % (v.year, v.month)
    return None


def _sheet(wb, name):
    n = _norm(name)
    for ws in wb.worksheets:
        if _norm(ws.title) == n:
            return ws
    raise ValueError("Sheet '%s' not found — is this the Modelo Polaris workbook?" % name)


def _month_header(ws, row=1) -> dict:
    """col index -> 'YYYY-MM' for every datetime cell in the header row."""
    out = {}
    for c in range(1, ws.max_column + 1):
        ym = _ym(ws.cell(row=row, column=c).value)
        if ym:
            out[c] = ym
    if len(out) < 24:
        raise ValueError("'%s' has no monthly date axis in row %d." % (ws.title, row))
    return out


def _label_rows(ws, col=2):
    """(row, normalised label, raw label) for every non-empty label cell."""
    out = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        n = _norm(v)
        if n:
            out.append((r, n, str(v).strip()))
    return out


# ── revenue blocks ───────────────────────────────────────────────────────────
def _revenue_totals(ws) -> dict:
    """Highway model name -> row index of its 'Total' revenue row, taken from
    the 'Ingresos' label block (the last such header before the annual
    rollup)."""
    labels = _label_rows(ws)
    start = None
    for r, n, _raw in labels:
        if n == "ingresos":
            start = r            # keep the LAST plain 'Ingresos' header
    if start is None:
        raise ValueError("'%s': no 'Ingresos' block found." % ws.title)
    rows, cur = {}, None
    for r, n, raw in labels:
        if r <= start:
            continue
        if n in ("tpda total", "ingresos anuales", "ingresos totales"):
            break
        if n == "total":
            if cur:
                rows[cur] = r
                cur = None
        elif n not in _CLASS_BY_NORM:
            cur = raw
    if not rows:
        raise ValueError("'%s': no highway Total rows under 'Ingresos'." % ws.title)
    return rows


def _class_block(ws, header_norm: str, stop_norms: set) -> dict:
    """For a block that begins at the label `header_norm` (e.g. 'trafico',
    'tarifas'): highway model name -> {class -> row}."""
    labels = _label_rows(ws)
    start = None
    for r, n, _raw in labels:
        if n == header_norm:
            start = r
            break
    if start is None:
        return {}
    out, cur = {}, None
    for r, n, raw in labels:
        if r <= start:
            continue
        if n in stop_norms:
            break
        if n in _CLASS_BY_NORM:
            if cur:
                out.setdefault(cur, {})[_CLASS_BY_NORM[n]] = r
        elif n == "total":
            continue
        elif n.startswith("incremento") or n.startswith("actualizacion") or n.startswith("estudio"):
            continue
        else:
            cur = raw
    return out


def _series(ws, row, hdr: dict, axis: list) -> list:
    by_ym = {}
    for c, ym in hdr.items():
        by_ym[ym] = by_ym.get(ym, 0.0) + _num(ws.cell(row=row, column=c).value)
    return [round(by_ym.get(ym, 0.0), 2) for ym in axis]


def _value_at(ws, row, hdr: dict, ym: str):
    for c, m in hdr.items():
        if m == ym:
            return _num(ws.cell(row=row, column=c).value)
    return None


# ── inputs block (Saldos Iniciales) ──────────────────────────────────────────
def _saldos_inputs(ws) -> dict:
    """'Datos de Ingresos' block: highway -> {tpda:{cls}, tarifa_con_iva:{cls},
    tarifa_sin_iva:{cls}}."""
    anchor = None
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 400)):
        for c in row:
            if _norm(c.value).startswith("datos de ingresos"):
                anchor = (c.row, c.column)
                break
        if anchor:
            break
    if not anchor:
        return {}
    r0, c0 = anchor
    cols = {}
    for c in range(c0, c0 + 24):
        h = _norm(ws.cell(row=r0, column=c).value)
        if h == "tpda":
            cols["tpda"] = c
        elif "con iva" in h:
            cols["con"] = c
        elif "sin iva" in h:
            cols["sin"] = c
    out, cur = {}, None
    for r in range(r0 + 1, min(r0 + 120, ws.max_row + 1)):
        raw = ws.cell(row=r, column=c0).value
        n = _norm(raw)
        if not n:
            continue
        if n.startswith("calendario") or n == "promav" or n.startswith("activo"):
            break
        if n in _CLASS_BY_NORM:
            if cur:
                cls = _CLASS_BY_NORM[n]
                if "tpda" in cols:
                    out[cur]["tpda"][cls] = _num(ws.cell(row=r, column=cols["tpda"]).value)
                if "con" in cols:
                    out[cur]["tarifa_con_iva"][cls] = _num(ws.cell(row=r, column=cols["con"]).value)
                if "sin" in cols:
                    out[cur]["tarifa_sin_iva"][cls] = _num(ws.cell(row=r, column=cols["sin"]).value)
        elif n == "total":
            if cur and "tpda" in cols:
                out[cur]["tpda"]["Total"] = _num(ws.cell(row=r, column=cols["tpda"]).value)
        else:
            cur = str(raw).strip()
            out[cur] = {"tpda": {}, "tarifa_con_iva": {}, "tarifa_sin_iva": {}}
    return out


# ── premisas / portada / margins ─────────────────────────────────────────────
def _premisas(ws) -> dict:
    out = {}
    for r, n, _raw in _label_rows(ws):
        v = ws.cell(row=r, column=3).value
        if n.startswith("inicio de proyeccion"):
            out["model_start"] = _ym(v)
        elif n.startswith("fin de concesion meta"):
            out["concession_end_META"] = _ym(v)
        elif n.startswith("fin de concesion ipc"):
            out["concession_end_IPC"] = _ym(v)
        elif n.startswith("pesos reales al"):
            out["real_pesos_as_of"] = _ym(v)
    return out


def _discount_rate(ws) -> float | None:
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200)):
        for c in row:
            if _norm(c.value) in ("t.d real", "td real", "tasa de descuento real"):
                for dr, dc in ((1, 0), (0, 1), (1, 1), (0, -1)):
                    v = ws.cell(row=c.row + dr, column=max(1, c.column + dc)).value
                    if isinstance(v, (int, float)) and 0 < v < 1:
                        return float(v)
    return None


def _margin(ws) -> float | None:
    """Σ EBITDA / Σ Ingresos over the years with revenue, from an annual
    Flujo sheet whose row 1 carries the year axis."""
    years = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        y = None
        if isinstance(v, (int, float)) and 2000 < v < 2200:
            y = int(v)
        elif isinstance(v, (dt.datetime, dt.date)):
            y = v.year
        if y:
            years[c] = y
    if len(years) < 5:
        return None
    rev_row = eb_row = None
    for r, n, _raw in _label_rows(ws):
        if r < 3:
            continue
        if rev_row is None and n == "ingresos":
            rev_row = r
        elif eb_row is None and n == "ebitda":
            eb_row = r
    if not (rev_row and eb_row):
        return None
    R = E = 0.0
    for c in years:
        rv = _num(ws.cell(row=rev_row, column=c).value)
        if rv > 0:
            R += rv
            E += _num(ws.cell(row=eb_row, column=c).value)
    return round(E / R, 4) if R > 0 else None


# ── main ─────────────────────────────────────────────────────────────────────
def parse_polaris(file_bytes: bytes, filename: str = "") -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    try:
        ws_meta = _sheet(wb, "Ingresos META")
        ws_ipc = _sheet(wb, "Ingresos IPC")
    except ValueError as e:
        raise ValueError(str(e) + " Expected the Modelo Polaris workbook with "
                         "'Ingresos META' and 'Ingresos IPC' sheets.")

    sheets = [("META", ws_meta), ("IPC", ws_ipc)]
    hdrs = {co: _month_header(ws) for co, ws in sheets}

    # Canonical monthly axis: model start .. last month any highway earns.
    all_months = sorted(set(m for h in hdrs.values() for m in h.values()))
    totals = {co: _revenue_totals(ws) for co, ws in sheets}
    last_idx = 0
    raw_series = {}
    for co, ws in sheets:
        for name, row in totals[co].items():
            s = _series(ws, row, hdrs[co], all_months)
            raw_series[(co, name)] = s
            nz = [i for i, v in enumerate(s) if v]
            if nz:
                last_idx = max(last_idx, nz[-1])
    months = all_months[: last_idx + 1]
    if len(months) < 12:
        raise ValueError("Parsed fewer than 12 months of revenue — wrong workbook?")
    years = sorted(set(int(m[:4]) for m in months))

    # Inputs + assumptions
    saldos = {}
    try:
        saldos = _saldos_inputs(_sheet(wb, "Saldos Iniciales"))
    except ValueError:
        pass
    saldos_by_key = {slug(k): v for k, v in saldos.items()}
    prem = {}
    try:
        prem = _premisas(_sheet(wb, "Premisas"))
    except ValueError:
        pass
    disc = None
    try:
        disc = _discount_rate(_sheet(wb, "Portada"))
    except ValueError:
        pass
    margins = {}
    for co, sh in (("META", "Flujo META Anual"), ("IPC", "Flujo IPC Anual")):
        try:
            margins[co] = _margin(_sheet(wb, sh))
        except ValueError:
            margins[co] = None

    # Trafico / Tarifas blocks (used at opening month for roads not in Saldos)
    traf = {co: _class_block(ws, "trafico", {"tarifas"}) for co, ws in sheets}
    tarf = {co: _class_block(ws, "tarifas", {"ingresos"}) for co, ws in sheets}

    highways = []
    order = 0
    for co, ws in sheets:
        for name, row in totals[co].items():
            s = raw_series[(co, name)][: len(months)]
            nz = [i for i, v in enumerate(s) if v]
            open_idx = nz[0] if nz else None
            key = slug(name)
            annual = {}
            for m, v in zip(months, s):
                annual[m[:4]] = round(annual.get(m[:4], 0.0) + v, 2)
            inp = saldos_by_key.get(key) or {}
            tpda = dict(inp.get("tpda") or {})
            tar_sin = dict(inp.get("tarifa_sin_iva") or {})
            tar_con = dict(inp.get("tarifa_con_iva") or {})
            tpda_src = "saldos" if tpda.get("Total") else None
            # Fall back to the model's own Trafico / Tarifas at opening month
            if open_idx is not None:
                om = months[open_idx]
                if not tpda.get("Total"):
                    trows = None
                    for k2, v2 in traf[co].items():
                        if slug(k2) == key:
                            trows = v2
                    if trows:
                        tot = 0.0
                        for cls, r in trows.items():
                            v = _value_at(ws, r, hdrs[co], om)
                            if v is not None:
                                tpda[cls] = round(v, 1)
                                tot += v
                        if tot:
                            tpda["Total"] = round(tot, 1)
                            tpda_src = "model@open"
                if not tar_sin:
                    trows = None
                    for k2, v2 in tarf[co].items():
                        if slug(k2) == key:
                            trows = v2
                    if trows:
                        for cls, r in trows.items():
                            v = _value_at(ws, r, hdrs[co], om)
                            if v:
                                tar_sin[cls] = round(v, 2)
            # TPDA-weighted average tarifa (sin IVA)
            avg_tarifa = None
            w = sum(tpda.get(c, 0) for c in CLASSES)
            if w and tar_sin:
                avg_tarifa = round(sum(tpda.get(c, 0) * tar_sin.get(c, 0) for c in CLASSES) / w, 2)
            order += 1
            highways.append(dict(
                key=key, model_name=name, company=co, sort=order,
                open_month=(months[open_idx] if open_idx is not None else None),
                open_idx=open_idx,
                in_development=bool(open_idx),          # opens after model start
                monthly=s, annual=annual,
                lifetime=round(sum(s), 2),
                margin=margins.get(co),
                tpda=tpda, tpda_src=tpda_src,
                tarifa_sin_iva=tar_sin, tarifa_con_iva=tar_con,
                avg_tarifa_sin_iva=avg_tarifa,
            ))

    # Cross-check against the sheets' own lifetime totals (col C of the header block)
    checks = {}
    for co, ws in sheets:
        for r, n, raw in _label_rows(ws):
            if r > 20:
                break
            if n in ("ingresos totales",):
                checks[co] = _num(ws.cell(row=r, column=3).value)
    parsed_tot = {co: round(sum(h["lifetime"] for h in highways if h["company"] == co), 2) for co in ("META", "IPC")}
    for co, sheet_tot in checks.items():
        if sheet_tot and abs(parsed_tot[co] - sheet_tot) > 0.001 * sheet_tot:
            raise ValueError("%s revenue parsed %.2f bn but the sheet says %.2f bn — layout changed?"
                             % (co, parsed_tot[co] / 1e9, sheet_tot / 1e9))

    by_year_total = {str(y): 0.0 for y in years}
    for h in highways:
        for y, v in h["annual"].items():
            by_year_total[y] = round(by_year_total.get(y, 0.0) + v, 2)

    return {
        "meta": {
            "filename": filename,
            "parsed_at": dt.datetime.utcnow().isoformat(timespec="seconds") + "Z",
            "model_start": prem.get("model_start") or months[0],
            "horizon_end": months[-1],
            "concession_end": {"META": prem.get("concession_end_META"),
                               "IPC": prem.get("concession_end_IPC")},
            "real_pesos_as_of": prem.get("real_pesos_as_of"),
            "discount_rate": disc if disc is not None else 0.11,
            "discount_rate_src": "Portada T.D Real" if disc is not None else "default",
            "margins": margins,
            "currency": "MXN", "iva": "sin IVA", "basis": "nominal",
            "sheet_totals": checks,
        },
        "months": months,
        "years": years,
        "highways": highways,
        "totals": {"lifetime": round(sum(h["lifetime"] for h in highways), 2),
                   "by_company": parsed_tot, "by_year": by_year_total},
    }
